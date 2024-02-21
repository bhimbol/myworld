from django import forms
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect
from django.template import loader
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from .models import Product, Promo, Valuation
import requests
import pandas as pd
import json
import os.path
import os
from django.views.decorators.http import require_POST
import ast
import logging
from django.conf import settings
import shutil
import base64
import xlrd
from openpyxl import load_workbook



logger = logging.getLogger(__name__)

def render_template(request, template_name, context=None):
    context = context or {}
    template = loader.get_template(template_name)
    request.session.save()
    return HttpResponse(template.render(context, request))

def dashboard(request):
    return render_template(request, 'dashboard.html')

def about(request):
    return render_template(request, 'about.html')

def product_details(request):
    try:
        products = Product.objects.all().order_by('description')
        context = {'products_list': products}
        return render(request, 'product_details.html', context)
    except Exception as e:
        return render(request, 'product_details.html', {'error_message': str(e)})

def promo_calculator(request):
    result_list = []
    num_of_deals = 0
    promo_description = ""
    if request.method == 'POST':
        promo_code = request.POST.get('promo', 0)
        promo_description = Product.objects.filter(sku=promo_code).first().description if Product.objects.filter(sku=promo_code).first() else None
        promos_product = Promo.objects.filter(sku=promo_code)
        num_of_deals = int(request.POST.get('num_of_deals', 0))
        for promo in promos_product:
            product = Product.objects.filter(sku=promo.child_sku).first()

            if product:
                product_description = product.description
                conv_factor = product.qtyperpcs
            else:
                product_description = "error"
                conv_factor = "error"

            qty_pcs = (promo.mech_per_pcs * num_of_deals) % conv_factor
            qty_cs = (((promo.mech_per_pcs * num_of_deals) - qty_pcs) / conv_factor)

            if qty_pcs == 0:
                qty_pcs = ""

            if qty_cs == 0:
                qty_cs = ""

            result_list.append({
                'child_sku': promo.child_sku,
                'mech_per_pcs': promo.mech_per_pcs,
                'description': product_description,
                'conv_factor': conv_factor,
                'qty_cs': qty_cs,
                'qty_pcs': qty_pcs
            })
    return render(request, 'promo_calc.html', {
        'promo_description': promo_description,
        'num_of_deals': num_of_deals,
        'result_list': result_list,
    })

def delete_records_with_condition():
    Product.objects.filter(sku__contains='(').delete()

def process_excel_file(uploaded_file):
    updated_cnt = 0
    added_cnt = 0
    try:
        df = pd.read_excel(uploaded_file)
        for index, row in df.iterrows():
            sku=row['SKU']
            existing_valuation = Product.objects.filter(sku=sku).first()
            if existing_valuation:
                if pd.notna(row['BCCS']) or pd.notna(row['BCIB']) or pd.notna(row['BCPCS']):
                    if pd.notna(row['BCCS']) and row['BCCS'] != existing_valuation.bccs:
                        existing_valuation.bccs = row['BCCS']
                    if pd.notna(row['BCIB']) and row['BCIB'] != existing_valuation.bcib:
                        existing_valuation.bcib = row['BCIB']
                    if pd.notna(row['BCPCS']) and row['BCPCS'] != existing_valuation.bcpcs:
                        existing_valuation.bcpcs = row['BCPCS']
                    updated_cnt = updated_cnt + 1
                    existing_valuation.save()
            else:
                Product.objects.create(
                    sku=sku,
                    description=row['DESCRIPTION'],
                    qtyperpcs=row['QTYPERPCS'],
                    bccs = row['BCCS'],
                    bcib = row['BCIB'],
                    bcpcs = row['BCPCS'])
                added_cnt = added_cnt + 1
        error_message = "UPDATED SKU: " + str(updated_cnt) + " ; ADDED SKU: " +  str(added_cnt)
    except Exception as e:
        error_message = f"An exception occurred: {str(e)}"
    return error_message

class FileUploadForm(forms.Form):
    file = forms.FileField()
    number_field = forms.IntegerField(label='Secret Key')

def upload_pd(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid() and form.cleaned_data['number_field'] == 8888:
            uploaded_file = request.FILES['file']
            error_message = process_excel_file(uploaded_file)
            context = {'error_message': error_message}
            template = loader.get_template('dashboard.html')
            return HttpResponse(template.render(context))
    else:
        form = FileUploadForm()
    return render(request, 'upload_pd.html', {'form': form})

def val(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid() and form.cleaned_data['number_field'] == 8888:
            try:
                uploaded_file = request.FILES['file']
                Valuation.objects.all().delete()
                df = pd.read_excel(uploaded_file)
                for index, row in df.iterrows():
                    Valuation.objects.create(
                        sku=row['StockCode'],
                        description=row['Description'],
                        cs=row['CS'],
                        ib = row['IB'],
                        pcs = row['PCS'])
                return HttpResponse("Data transfer successful")
            except Exception as e:
                return HttpResponse(f"{e}")
    else:
        form = FileUploadForm()
    return render(request, 'val.html', {'form': form})

def get_path_from_static(fname):
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    DJANGO_PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, '..'))
    STATIC_FOLDER = os.path.join(DJANGO_PROJECT_ROOT, 'static')
    path = os.path.join(STATIC_FOLDER, fname)
    return path

def revoke_google_credentials(credentials):
    revoke_url = "https://accounts.google.com/o/oauth2/revoke"
    revoke_params = {"token": credentials.token}
    requests.post(revoke_url, params=revoke_params)

def asn_logout(request):
    try:
        if 'google_credentials' in request.session:
            google_credentials_json = request.session.get('google_credentials')
            credentials = Credentials.from_authorized_user_info(json.loads(google_credentials_json))
            revoke_google_credentials(credentials)
            del request.session['user_info']
            del request.session['google_credentials']
    except Exception as e:
        return HttpResponse(f"Error: {e}")
    return redirect('asn')

def initiate_google_auth(request):
    credentials_path = get_path_from_static("keyMe.json")
    with open(credentials_path, 'r') as file:
        credentials_data = json.load(file)
    flow = InstalledAppFlow.from_client_config(
        credentials_data,
        scopes=['https://www.googleapis.com/auth/gmail.readonly'],
        redirect_uri='https://airbytes.pythonanywhere.com/callback/'
    )
    authorization_url, state = flow.authorization_url(prompt="consent")
    request.session['google_auth_state'] = state
    return HttpResponseRedirect(authorization_url)

def callback_view(request):
    authorization_code = request.GET.get('code')
    received_state = request.GET.get('state')
    expected_state = request.session.pop('google_auth_state', None)

    if received_state != expected_state:
        return HttpResponse(f"Invalid state. Possible CSRF attack.\nReceived State: {received_state}\nExpected State: {expected_state}")

    try:
        credentials_path = get_path_from_static("keyMe.json")
        with open(credentials_path, 'r') as file:
            credentials_data = json.load(file)

        flow = InstalledAppFlow.from_client_config(
            credentials_data,
            scopes=['https://www.googleapis.com/auth/gmail.readonly'],
            redirect_uri='https://airbytes.pythonanywhere.com/callback/'
        )

        flow.fetch_token(code=authorization_code)
        credentials = flow.credentials

        gmail_service = build('gmail', 'v1', credentials=credentials)
        user_info = gmail_service.users().getProfile(userId='me').execute()

        request.session['user_info'] = user_info
        request.session['google_credentials'] = credentials.to_json()

        return redirect('/asn')
    except Exception as e:
        return HttpResponse(f"Error {e}")

def get_gmail_messages(request, keywords=None, rdd=None, sender=None, has_attachment=True):
    try:
        if 'google_credentials' in request.session:
            google_credentials_json = request.session.get('google_credentials')
            credentials = Credentials.from_authorized_user_info(json.loads(google_credentials_json))
            gmail_service = build('gmail', 'v1', credentials=credentials)
            query_parts = []
            if has_attachment:
                query_parts.append('has:attachment')
            if sender:
                query_parts.append(f'from:{sender}')
            if keywords:
                for keyword in keywords:
                    query_parts.append(f'subject:{keyword}')
            query = ' '.join(query_parts)

            messages = gmail_service.users().messages().list(userId='me', q=query).execute()
            message_list = []
            for message in messages.get('messages', []):
                message_details = gmail_service.users().messages().get(userId='me', id=message['id']).execute()
                subject = next((header['value'] for header in message_details['payload']['headers'] if header['name'] == 'Subject'), 'N/A')
                attachments = []
                if 'parts' in message_details['payload']:
                    for part in message_details['payload']['parts']:
                        if 'data' in part['body']:
                            data = part['body']['data']
                        elif 'attachmentId' in part['body']:
                            att_id = part['body']['attachmentId']
                            att = gmail_service.users().messages().attachments().get(userId='me', messageId=message['id'], id=att_id).execute()
                            data = att['data']
                        file_data = base64.urlsafe_b64decode(data.encode('UTF-8'))
                        #path = part['filename']
                        attachments.append({'filename': part['filename'], 'data': file_data})
                message_list.append({'id': message['id'], 'subject': subject, 'attachments': attachments})
            return message_list
    except Exception as e:
        print(f"Error: {e}")
        return []

class ASNForm(forms.Form):
    rdd = forms.DateField(
        label='Request Delivery Date',
        widget=forms.TextInput(attrs={'type': 'date'}),
        input_formats=['%Y-%m-%d']
    )

def asn(request):
    form = ASNForm()
    if 'google_credentials' in request.session:
        google_authenticated = True
    else:
        google_authenticated = False

    if request.method == 'POST':
        form = ASNForm(request.POST)
        if form.is_valid():
            rdd = form.cleaned_data['rdd']
            try:
                messages = []
                messages = get_gmail_messages(request, keywords=['FDC-Borongan, PO#:', rdd.strftime('%d %b %Y')], sender='kmgflores@fastgroup.biz', has_attachment=True)
                messages_count = len(messages)
            except Exception as e:
                return JsonResponse({'error': f"Error: {e}"})
    return render(request, 'asn.html', {'form': form, 'google_authenticated': google_authenticated, 'messages': messages, 'messages_count': messages_count})

@require_POST
def download_selected_attachments(request):
    selected_messages_list = request.POST.getlist('selected_messages')
    to_download_list = []
    save_dir = os.path.join(settings.BASE_DIR, 'attachments', '')

    if not selected_messages_list:
        return JsonResponse({'to_download_list': []})

    if os.path.exists(save_dir):
        shutil.rmtree(save_dir)
    os.makedirs(save_dir)

    try:
        for message_str in selected_messages_list:
            message = ast.literal_eval(message_str)
            attachments = message.get('attachments', [])

            for attachment in attachments:
                filename = attachment.get('filename', '')
                file_path = os.path.join(save_dir, filename)
                data = attachment.get('data', '')
                if '.xls' in filename:
                    with open(file_path, 'wb') as file:
                        file.write(data)
                    to_download_list.append({'file_path':file_path})
        return HttpResponse("Successfully Downloaded ASN Files. Proceed to <a href='/asn_process/'>Process ASN</a>.")
    except (SyntaxError, ValueError) as e:
        return HttpResponse(f"{e}")

def asn_process(request):
    existing_file_name = 'summary.xlsx'
    existing_file_path = os.path.join(settings.BASE_DIR, 'static', existing_file_name)

    asn_list = os.path.join(settings.BASE_DIR, 'attachments')
    asn_list_file_paths = [os.path.join(asn_list, filename) for filename in os.listdir(asn_list) if filename.endswith('.xls')]

    response = append_contents_to_existing_final(existing_file_path, asn_list_file_paths)
    return response

def append_contents_to_existing_final(existing_final_file_path, list_of_files):
    final_workbook = load_workbook(existing_final_file_path)

    final_sheet = final_workbook['DATA']
    final_sheet.delete_rows(1, final_sheet.max_row)

    for file_path in list_of_files:
        source_workbook = xlrd.open_workbook(file_path)
        source_sheet = source_workbook.sheet_by_index(0)

        for row_index in range(source_sheet.nrows):
            new_row = source_sheet.row_values(row_index)
            final_sheet.append(new_row)

    data_column_i_values = [final_sheet.cell(row=row_index, column=9).value for row_index in range(1, final_sheet.max_row + 1)]
    comparison_sheet = final_workbook['COMPARISON']

    for row in comparison_sheet.iter_rows(min_row=2, max_row=comparison_sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.value = None

    unique_values = list(set(data_column_i_values))
    unique_values = [value for value in unique_values if value is not None and pd.notna(value)]

    start_row = 2

    for value in unique_values:
        comparison_sheet.cell(row=start_row, column=1, value=value)
        start_row += 1

    final_workbook.save(existing_final_file_path)

    with open(existing_final_file_path, 'rb') as file:
        response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={existing_final_file_path}'

    return response